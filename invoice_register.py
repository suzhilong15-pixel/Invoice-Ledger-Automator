from __future__ import annotations

import argparse
import re
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "invoices_to_process"
ARCHIVE_DIR = BASE_DIR / "archived_invoices"
FAILED_DIR = BASE_DIR / "failed_invoices"
LEDGER_PATH = BASE_DIR / "invoice_ledger.xlsx"

HEADERS = [
    "发票号码",
    "开票日期",
    "购买方名称",
    "销售方名称",
    "价税合计（小写）",
    "美金金额",
    "汇率",
    "提单号",
    "源文件名",
    "识别状态",
    "失败原因",
    "处理时间",
]

REQUIRED_FIELDS = ["发票号码", "开票日期", "购买方名称", "销售方名称", "价税合计（小写）"]
SEPARATORS = r"[\s:：,，;；、为]*"


@dataclass
class InvoiceRecord:
    invoice_number: str = ""
    invoice_date: str = ""
    buyer_name: str = ""
    seller_name: str = ""
    total_amount: str = ""
    usd_amount: str = "无"
    exchange_rate: str = "无"
    bill_of_lading: str = "无"
    source_file: str = ""
    status: str = "成功"
    failed_reason: str = ""
    processed_at: str = field(default_factory=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    def as_row(self) -> list[str]:
        return [
            self.invoice_number,
            self.invoice_date,
            self.buyer_name,
            self.seller_name,
            self.total_amount,
            self.usd_amount,
            self.exchange_rate,
            self.bill_of_lading,
            self.source_file,
            self.status,
            self.failed_reason,
            self.processed_at,
        ]


def extract_pdf_text(pdf_path: Path) -> str:
    with pdfplumber.open(pdf_path) as pdf:
        pages = [page.extract_text(x_tolerance=1, y_tolerance=3) or "" for page in pdf.pages]
    return "\n".join(pages)


def compact_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def first_match(patterns: list[str], text: str, flags: int = re.IGNORECASE) -> Optional[str]:
    for pattern in patterns:
        match = re.search(pattern, text, flags)
        if match:
            value = match.group(1).strip()
            if value:
                return value
    return None


def normalize_date(value: str) -> str:
    value = value.strip()
    match = re.search(r"(\d{4})\s*[年./-]?\s*(\d{1,2})\s*[月./-]?\s*(\d{1,2})", value)
    if not match:
        return value
    year, month, day = match.groups()
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def normalize_amount(value: str) -> str:
    return value.replace(",", "").replace("，", "").strip()


def extract_invoice_number(text: str, one_line: str) -> Optional[str]:
    value = first_match(
        [
            rf"发\s*票\s*号码{SEPARATORS}([A-Z0-9]{{6,30}})",
            rf"发票号码{SEPARATORS}([A-Z0-9]{{6,30}})",
            rf"发票号码{SEPARATORS}\n?{SEPARATORS}([A-Z0-9]{{6,30}})",
        ],
        text,
    )
    if value:
        return value

    # 老版发票的 PDF 文本可能把“发票号码”和号码分散在版面不同位置。
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for idx, line in enumerate(lines):
        if "发票代码" in line and idx + 1 < len(lines):
            candidate = re.search(r"\b(\d{8})\b", lines[idx + 1])
            if candidate:
                return candidate.group(1)

    title_match = re.search(r"发\s*票\s*监\s+(\d{8,20})", one_line)
    if title_match:
        return title_match.group(1)

    legacy_match = re.search(r"电子普通发票\s+统一\s+发\s+票监\s+(\d{8,20})", one_line)
    if legacy_match:
        return legacy_match.group(1)
    return None


def extract_invoice_date(text: str, one_line: str) -> Optional[str]:
    value = first_match(
        [
            rf"开票\s*日期{SEPARATORS}(\d{{4}}\s*[年./-]\s*\d{{1,2}}\s*[月./-]\s*\d{{1,2}}\s*日?)",
            rf"开票日期{SEPARATORS}(\d{{4}}\s*[年./-]\s*\d{{1,2}}\s*[月./-]\s*\d{{1,2}}\s*日?)",
            rf"开票日期{SEPARATORS}(\d{{4}}\s+\d{{1,2}}\s+\d{{1,2}})",
        ],
        text,
    )
    if value:
        return normalize_date(value)

    title_match = re.search(r"国家税务总局\s+章\s+(\d{4})\s+(\d{1,2})\s+(\d{1,2})", one_line)
    if not title_match:
        title_match = re.search(r"税\s*务\s*总局\s+章\s+(\d{4})\s+(\d{1,2})\s+(\d{1,2})", one_line)
    if title_match:
        return normalize_date(" ".join(title_match.groups()))
    return None


def extract_names(text: str, one_line: str) -> tuple[Optional[str], Optional[str]]:
    buyer = None
    seller = None

    combined = first_match(
        [
            rf"购\s*名\s*称{SEPARATORS}(.+?)\s+销\s*名\s*称{SEPARATORS}(.+?)(?:\s+买|\s+售|\s+方|\s+信|\s+统一社会信用代码)",
            rf"购买方名\s*称{SEPARATORS}(.+?)\s+销售方名\s*称{SEPARATORS}(.+?)(?:\s+统一社会信用代码|\s+项目名称)",
        ],
        one_line,
    )
    if combined:
        # first_match only returns group 1, so use a direct search when this format exists.
        match = re.search(
            rf"购\s*名\s*称{SEPARATORS}(.+?)\s+销\s*名\s*称{SEPARATORS}(.+?)(?:\s+买|\s+售|\s+方|\s+信|\s+统一社会信用代码)",
            one_line,
        )
        if match:
            buyer = clean_name(match.group(1))
            seller = clean_name(match.group(2))

    if not buyer or not seller:
        match = re.search(
            rf"购\s*名\s*称{SEPARATORS}(.+?)\s+销\s*名\s*称{SEPARATORS}(.+?)\s+(?:买|售|方|信|统一社会信用代码)",
            one_line,
        )
        if match:
            buyer = buyer or clean_name(match.group(1))
            seller = seller or clean_name(match.group(2))

    if not buyer or not seller:
        for line in text.splitlines():
            match = re.search(rf"名\s*称{SEPARATORS}(.+?)\s+名\s*称{SEPARATORS}(.+)$", line.strip())
            if not match:
                continue
            candidate_buyer = clean_name(match.group(1))
            candidate_seller = clean_name(match.group(2))
            if "公司" in candidate_buyer and "公司" in candidate_seller:
                buyer = buyer or candidate_buyer
                seller = seller or candidate_seller
                break

    if not buyer:
        buyer = extract_legacy_buyer(text)
    if not seller:
        seller = extract_legacy_seller(text)

    return buyer, seller


def clean_name(value: str) -> str:
    value = re.sub(r"\s+", "", value)
    value = re.sub(r"^(名称|名 称|名\s*称)[:：]?", "", value)
    value = re.sub(r"^(：|:)", "", value)
    return value.strip(" :：,，;；")


def extract_legacy_buyer(text: str) -> Optional[str]:
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.endswith("有限公司") and " " not in stripped and len(stripped) >= 6:
            return stripped
        match = re.match(r"^([\u4e00-\u9fa5（）()A-Za-z0-9]+有限公司)\b", stripped)
        if match:
            return match.group(1)
    return None


def extract_legacy_seller(text: str) -> Optional[str]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    company_candidates: list[str] = []
    for idx, line in enumerate(lines):
        if "USD" in line.upper() and "有限公司" in line:
            before_usd = re.split(r"USD", line, flags=re.IGNORECASE)[0]
            candidates = re.findall(r"([\u4e00-\u9fa5（）()A-Za-z0-9]+有限公司)", before_usd)
            if candidates:
                return candidates[-1]

        company_candidates.extend(re.findall(r"([\u4e00-\u9fa5（）()A-Za-z0-9]+有限公司)", line))

        if re.fullmatch(r"[\u4e00-\u9fa5（）()A-Za-z0-9]+有限公司", line):
            nearby = "\n".join(lines[idx : idx + 5])
            if "开户行" in nearby or "银行" in nearby or "USD" in nearby.upper():
                return line
    if len(company_candidates) >= 2:
        return company_candidates[-1]
    return None


def extract_total_amount(text: str, one_line: str) -> Optional[str]:
    value = first_match(
        [
            rf"[（(]\s*小\s*写\s*[）)]{SEPARATORS}[¥￥]?\s*([0-9][0-9,，]*\.?\d*)",
            rf"价税合计.*?[（(]\s*小\s*写\s*[）)]{SEPARATORS}[¥￥]?\s*([0-9][0-9,，]*\.?\d*)",
        ],
        one_line,
    )
    if value:
        return normalize_amount(value)

    legacy_total = re.search(r"[壹贰叁肆伍陆柒捌玖拾佰仟万亿零圆元角分整]+\s*[¥￥]\s*([0-9][0-9,，]*\.?\d*)", one_line)
    if legacy_total:
        return normalize_amount(legacy_total.group(1))

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    for idx, line in enumerate(lines):
        if "价税合计" in line and "小写" in line:
            window = " ".join(lines[idx : idx + 3])
            amounts = re.findall(r"[¥￥]\s*([0-9][0-9,，]*\.?\d*)", window)
            if amounts:
                return normalize_amount(amounts[-1])
    return None


def extract_keyword_number(keyword: str, text: str, kind: str = "amount") -> tuple[str, bool]:
    text = re.sub(r"USD\s*1\s*=", "", text, flags=re.IGNORECASE)
    if keyword.upper() == "USD":
        if "USD" not in text.upper():
            return "无", False
        pattern = rf"USD{SEPARATORS}([0-9][0-9,，]*\.?\d*)(?!\s*=)"
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return normalize_amount(match.group(1)), False
        return "识别失败", True

    if keyword not in text:
        return "无", False

    if kind == "code":
        pattern = rf"{keyword}{SEPARATORS}([A-Za-z0-9][A-Za-z0-9_\-\/]*)(?=[\s,，;；。]|$)"
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).strip(), False
        return "识别失败", True

    pattern = rf"{keyword}{SEPARATORS}([0-9][0-9,，]*\.?\d*)"
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        return normalize_amount(match.group(1)), False
    return "识别失败", True


def extract_remark_text(text: str) -> str:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    remark_lines: list[str] = []

    for idx, line in enumerate(lines):
        is_total_value_line = ("价税合计" in line and re.search(r"[¥￥]\s*\d", line)) or re.search(
            r"[壹贰叁肆伍陆柒捌玖拾佰仟万亿零圆元角分整]+\s*[¥￥]\s*\d", line
        )
        if is_total_value_line:
            window = lines[idx + 1 :]
            for candidate in window:
                if candidate.startswith("开票人") or "收款人:" in candidate or "收款人：" in candidate:
                    break
                remark_lines.append(candidate)
            break

    if not remark_lines:
        for idx, line in enumerate(lines):
            if line in {"备", "注"} or "备注" in line or "备 注" in line:
                remark_lines.extend(lines[idx + 1 : idx + 8])
                break

    clean_lines: list[str] = []
    for line in remark_lines:
        if re.search(r"开户行|银行账号|账号|账户", line):
            continue
        if re.fullmatch(r"\d{15,20}[A-Z0-9]*", line):
            break
        clean_lines.append(line)

    return compact_text("\n".join(clean_lines))


def parse_invoice(pdf_path: Path) -> InvoiceRecord:
    record = InvoiceRecord(source_file=pdf_path.name)
    failures: list[str] = []

    try:
        text = extract_pdf_text(pdf_path)
    except Exception as exc:
        record.status = "失败"
        record.failed_reason = f"PDF文本读取失败：{exc}"
        return record

    one_line = compact_text(text)
    remark_text = extract_remark_text(text)
    record.invoice_number = extract_invoice_number(text, one_line) or ""
    record.invoice_date = extract_invoice_date(text, one_line) or ""
    record.buyer_name, record.seller_name = extract_names(text, one_line)
    record.buyer_name = record.buyer_name or ""
    record.seller_name = record.seller_name or ""
    record.total_amount = extract_total_amount(text, one_line) or ""

    record.usd_amount, usd_failed = extract_keyword_number("USD", remark_text, kind="amount")
    record.exchange_rate, rate_failed = extract_keyword_number("汇率", remark_text, kind="amount")
    record.bill_of_lading, bol_failed = extract_keyword_number("提单号", remark_text, kind="code")

    values_by_field = {
        "发票号码": record.invoice_number,
        "开票日期": record.invoice_date,
        "购买方名称": record.buyer_name,
        "销售方名称": record.seller_name,
        "价税合计（小写）": record.total_amount,
    }
    for field_name in REQUIRED_FIELDS:
        if not values_by_field[field_name]:
            failures.append(f"{field_name}识别失败")

    if usd_failed:
        failures.append("检测到USD关键词，但美金金额识别失败")
    if rate_failed:
        failures.append("检测到汇率关键词，但汇率数字识别失败")
    if bol_failed:
        failures.append("检测到提单号关键词，但提单号识别失败")

    if failures:
        record.status = "失败"
        record.failed_reason = "；".join(failures)

    return record


def ensure_workbook(path: Path):
    if path.exists():
        workbook = load_workbook(path)
        sheet = workbook.active
        if [cell.value for cell in sheet[1]] != HEADERS:
            raise ValueError(f"{path.name} 的表头与脚本要求不一致，请检查后再运行。")
        return workbook, sheet

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoice Ledger"
    sheet.append(HEADERS)
    style_header(sheet)
    set_column_widths(sheet)
    sheet.freeze_panes = "A2"
    workbook.save(path)
    return workbook, sheet


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def set_column_widths(sheet) -> None:
    widths = [22, 14, 34, 34, 18, 16, 12, 24, 42, 12, 48, 20]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width


def value_score(value) -> float:
    if value is None:
        return 0
    value = str(value).strip()
    if not value or value == "识别失败":
        return 0
    if value == "无":
        return 0.2
    return 1


def row_completeness_score(row_values: list) -> float:
    score = 0.0
    if len(row_values) >= 10 and row_values[9] == "成功":
        score += 100
    for idx in range(8):
        if idx < len(row_values):
            score += value_score(row_values[idx])
    return score


def deduplicate_sheet(sheet) -> None:
    if sheet.max_row <= 2:
        return

    rows = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]
    best_by_invoice: dict[str, list] = {}
    order: list[str] = []
    rows_without_invoice: list[list] = []

    for row in rows:
        invoice_number = str(row[0]).strip() if row and row[0] is not None else ""
        if not invoice_number:
            rows_without_invoice.append(row)
            continue

        if invoice_number not in best_by_invoice:
            best_by_invoice[invoice_number] = row
            order.append(invoice_number)
            continue

        current = best_by_invoice[invoice_number]
        if row_completeness_score(row) >= row_completeness_score(current):
            best_by_invoice[invoice_number] = row

    merged_rows = [best_by_invoice[invoice_number] for invoice_number in order] + rows_without_invoice
    sheet.delete_rows(2, sheet.max_row - 1)
    for row in merged_rows:
        sheet.append(row)


def row_should_be_removed(row_values: list) -> bool:
    status = str(row_values[9]).strip() if len(row_values) > 9 and row_values[9] is not None else ""
    failed_reason = str(row_values[10]).strip() if len(row_values) > 10 and row_values[10] is not None else ""
    return status == "失败" and "发票号码识别失败" in failed_reason


def remove_failed_rows_without_invoice_number(sheet) -> int:
    if sheet.max_row <= 1:
        return 0

    removed_count = 0
    for row_idx in range(sheet.max_row, 1, -1):
        row_values = [cell.value for cell in sheet[row_idx]]
        if row_should_be_removed(row_values):
            sheet.delete_rows(row_idx, 1)
            removed_count += 1
    return removed_count


def style_data_rows(sheet) -> None:
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def append_record(record: InvoiceRecord, ledger_path: Path) -> None:
    workbook, sheet = ensure_workbook(ledger_path)
    remove_failed_rows_without_invoice_number(sheet)
    sheet.append(record.as_row())
    remove_failed_rows_without_invoice_number(sheet)
    deduplicate_sheet(sheet)
    style_data_rows(sheet)
    workbook.save(ledger_path)


def unique_destination(directory: Path, filename: str) -> Path:
    destination = directory / filename
    if not destination.exists():
        return destination

    stem = destination.stem
    suffix = destination.suffix
    counter = 1
    while True:
        candidate = directory / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def move_pdf(pdf_path: Path, record: InvoiceRecord) -> Path:
    target_dir = ARCHIVE_DIR if record.status == "成功" else FAILED_DIR
    target_dir.mkdir(parents=True, exist_ok=True)
    destination = unique_destination(target_dir, pdf_path.name)
    shutil.move(str(pdf_path), str(destination))
    return destination


def process_invoices(dry_run: bool = False) -> tuple[int, int]:
    for directory in [INPUT_DIR, ARCHIVE_DIR, FAILED_DIR]:
        directory.mkdir(parents=True, exist_ok=True)

    pdf_files = sorted(list(INPUT_DIR.glob("*.pdf")) + list(INPUT_DIR.glob("*.PDF")))
    if not pdf_files:
        print(f"未发现待识别PDF：{INPUT_DIR}")
        return 0, 0

    success_count = 0
    failed_count = 0
    for pdf_path in pdf_files:
        record = parse_invoice(pdf_path)
        if record.status == "成功":
            success_count += 1
        else:
            failed_count += 1

        print(f"[{record.status}] {pdf_path.name}")
        if record.failed_reason:
            print(f"  原因：{record.failed_reason}")

        if not dry_run:
            append_record(record, LEDGER_PATH)
            destination = move_pdf(pdf_path, record)
            print(f"  已移动到：{destination}")

    return success_count, failed_count


def main() -> None:
    parser = argparse.ArgumentParser(description="识别PDF发票并追加写入Excel台账。")
    parser.add_argument("--dry-run", action="store_true", help="只识别并打印结果，不写入Excel，也不移动PDF。")
    args = parser.parse_args()

    success_count, failed_count = process_invoices(dry_run=args.dry_run)
    print(f"处理完成：成功 {success_count} 个，失败 {failed_count} 个。")


if __name__ == "__main__":
    main()
